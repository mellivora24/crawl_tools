#!/usr/bin/env python3
"""
Script tạo file Excel template mẫu cho danh sách sản phẩm
"""

import pandas as pd
import os
from pathlib import Path

def create_excel_template():
    """Tạo file Excel template mẫu"""
    
    # Dữ liệu mẫu
    sample_data = {
        'url': [
            'https://example.com/product1',
            'https://example.com/product2',
            'https://example.com/product3',
            'https://shop.example.com/item1',
            'https://shop.example.com/item2'
        ],
        'name': [
            'Sản phẩm mẫu 1',
            'Sản phẩm mẫu 2', 
            'Sản phẩm mẫu 3',
            'Sản phẩm mẫu 4',
            'Sản phẩm mẫu 5'
        ],
        'category': [
            'Điện tử',
            'Thời trang',
            'Gia dụng',
            'Sách',
            'Thể thao'
        ],
        'notes': [
            'Sản phẩm điện tử cao cấp',
            'Quần áo thời trang nam',
            'Đồ gia dụng nhà bếp',
            'Sách kỹ năng sống',
            'Dụng cụ thể thao'
        ]
    }
    
    # Tạo DataFrame
    df = pd.DataFrame(sample_data)
    
    # Đường dẫn file template
    template_dir = Path(__file__).parent.parent / 'io' / 'templates'
    template_dir.mkdir(parents=True, exist_ok=True)
    
    template_file = template_dir / 'product_links_template.xlsx'
    
    # Tạo file Excel với định dạng đẹp
    with pd.ExcelWriter(template_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Products', index=False)
        
        # Lấy workbook và worksheet để định dạng
        workbook = writer.book
        worksheet = writer.sheets['Products']
        
        # Định dạng header
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Điều chỉnh độ rộng cột
        worksheet.column_dimensions['A'].width = 40  # URL
        worksheet.column_dimensions['B'].width = 25  # Name
        worksheet.column_dimensions['C'].width = 15  # Category
        worksheet.column_dimensions['D'].width = 30  # Notes
        
        # Thêm sheet hướng dẫn
        guide_sheet = workbook.create_sheet("Hướng dẫn", 0)
        
        guide_data = [
            ["HƯỚNG DẪN SỬ DỤNG TEMPLATE"],
            [""],
            ["CẤU TRÚC CỘT:"],
            ["- url: URL của sản phẩm (bắt buộc)"],
            ["- name: Tên sản phẩm (tùy chọn)"],
            ["- category: Danh mục sản phẩm (tùy chọn)"],
            ["- notes: Ghi chú thêm (tùy chọn)"],
            [""],
            ["LƯU Ý:"],
            ["1. Cột 'url' là bắt buộc và phải chứa URL hợp lệ"],
            ["2. Các cột khác có thể để trống"],
            ["3. Xóa các dòng mẫu trước khi sử dụng"],
            ["4. Đảm bảo URL có thể truy cập được"],
            [""],
            ["VÍ DỤ:"],
            ["https://shop.example.com/product1, Laptop Dell, Máy tính, Laptop gaming cao cấp"],
            ["https://shop.example.com/product2, Điện thoại iPhone, Điện tử, iPhone 15 Pro Max"]
        ]
        
        for row_idx, row_data in enumerate(guide_data, 1):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = guide_sheet.cell(row=row_idx, column=col_idx, value=cell_data)
                if row_idx == 1:  # Tiêu đề
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 9, 15]:  # Các tiêu đề phụ
                    cell.font = Font(bold=True)
        
        # Điều chỉnh độ rộng cột cho sheet hướng dẫn
        guide_sheet.column_dimensions['A'].width = 60
    
    print(f"Đã tạo file template: {template_file}")
    print("File template chứa:")
    print("- Sheet 'Hướng dẫn': Hướng dẫn sử dụng")
    print("- Sheet 'Products': Dữ liệu mẫu và cấu trúc cột")
    
    return template_file

if __name__ == "__main__":
    create_excel_template() 