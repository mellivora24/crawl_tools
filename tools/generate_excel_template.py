import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.builtins import output

from utils.resource_path import resource_path as path_to

OUTPUT_DIR = path_to("../io/input")

def generate_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "product_links_template"

    # Tiêu đề cột
    headers = ["STT", "Product URL", "Is Crawled", "Crawled Time", "Note"]
    ws.append(headers)

    # Style tiêu đề
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 50

    wb.save(os.path.join(OUTPUT_DIR, "product_links_template.xlsx"))

    output_file_path = os.path.join(OUTPUT_DIR, "product_links_template.xlsx")
    print(f"[v] Đã tạo file Excel mẫu tại {output_file_path}")

if __name__ == "__main__":
    generate_excel_template()
